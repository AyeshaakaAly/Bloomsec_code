import hashlib
import os
from Crypto.Random import get_random_bytes
from sklearn.feature_extraction.text import TfidfVectorizer
from Crypto.Cipher import AES
from openpyxl import Workbook
import math
import base64
import numpy as np
import openpyxl

def extract_keywords_from_file(file_path):
    # Read the content of the file
    try:
        with open(file_path, 'r', encoding='latin1') as f:
            content = f.read()  # Read the entire content of the file
    except Exception as e:
        print(f"Error reading file {file_path}: {e}")
        return []

    # Initialize TfidfVectorizer to extract keywords
    vectorizer = TfidfVectorizer(stop_words='english')

    # Apply the vectorizer to the content
    tfidf_matrix = vectorizer.fit_transform([content])  # The input needs to be a list of documents (even if it's one)

    # Get the feature names (keywords) and their corresponding TF-IDF scores
    feature_names = vectorizer.get_feature_names_out()
    tfidf_scores = tfidf_matrix.sum(axis=0).A1  # Sum the TF-IDF scores across the document (row)

    # Combine keywords and their scores into a list of tuples
    keyword_scores = list(zip(feature_names, tfidf_scores))

    # Sort the keywords by their TF-IDF score in descending order
    sorted_keywords = sorted(keyword_scores, key=lambda x: x[1], reverse=True)

    # Return the top N keywords
    return sorted_keywords




def encryptData(data, key):
    key = str(key)  # Ensure key is a string
    kbb = key.encode('utf-8')
    kbb = kbb[:32] if len(kbb) > 32 else kbb.ljust(32, b'\0')  # Ensure 32-byte key

    # Ensure the data is a string and then encode it to bytes
    data = str(data)  # Convert data to string if necessary
    data_bytes = data.encode('utf-8', errors='ignore')

    # Generate a random IV (Initialization Vector)
    iv = get_random_bytes(12)  # GCM mode typically uses 12-byte IV

    cipher = AES.new(kbb, AES.MODE_GCM, nonce=iv)

    # Encrypt the data
    encrypted_data, tag = cipher.encrypt_and_digest(data_bytes)

    # Combine the IV, encrypted data, and the tag for storage
    encrypted_data_base64 = base64.b64encode(iv + tag + encrypted_data).decode('utf-8')

    return encrypted_data_base64

def decryptData(encrypted_data_base64, key):
    key = str(key)  # Ensure key is a string
    kbb = key.encode('utf-8')
    kbb = kbb[:32] if len(kbb) > 32 else kbb.ljust(32, b'\0')  # Ensure 32-byte key

    # Decode the base64 encoded encrypted data
    encrypted_data_with_nonce_and_tag = base64.b64decode(encrypted_data_base64)

    # Extract the IV, authentication tag, and encrypted data
    iv = encrypted_data_with_nonce_and_tag[:12]
    tag = encrypted_data_with_nonce_and_tag[12:28]
    encrypted_data = encrypted_data_with_nonce_and_tag[28:]

    # Create the AES cipher in GCM mode using the same IV and tag
    cipher = AES.new(kbb, AES.MODE_GCM, nonce=iv)
    
    # Decrypt the data and verify its authenticity
    decrypted_data = cipher.decrypt_and_verify(encrypted_data, tag)

    # Convert the decrypted data back to a string
    decrypted_data_str = decrypted_data.decode('utf-8', errors='ignore')  # Handle potential decoding issues

    return decrypted_data_str



    
    #return round(k), m, p
    
    
    #print(f"Keywords: {n} BF Size: {m} k: {k}")
def computeIntHash(n,r):
    sha256_hash = hashlib.sha256(n.encode()).hexdigest()
    # Step 2: Convert the hash (hexadecimal) to an integer
    hash_integer = int(sha256_hash, 16)%r
    return sha256_hash,hash_integer

# Path to dataset
def setup_Phase(dataset_pathm,Kgroup):
    #dataset_path = r"maildir"  # Replace with your dataset directory
    if not os.path.exists(dataset_path):
        print(f"Path does not exist: {dataset_path}")
    else:
        print(f"Processing files in: {dataset_path}")
    # Traverse directory to read email files
    
    wb = Workbook()

    # Select the active sheet
    sheet = wb.active
    

    # Write headers for the columns
    sheet['A1'] = 'path'
    sheet['B1'] = 'BF'
    sheet['C1'] = 'e'
    sheet['D1'] = 'k'
    sheet['E1'] = 'm'
    idx=2
    #New Code
    keyword_files_map = {}

    # Traverse the directory to read files
    for root, dirs, files in os.walk(dataset_path):
        for file in files:
            file_path = os.path.join(root, file)
            print(file_path)
            # Extract keywords from the current file
            keywords = extract_keywords_from_file(file_path)
            n=len(keywords)
            k,m,p=findBloomsParas(n)
            print(f"n: {n}, k: {k}, m: {m}, p: {p}")
            bf = np.zeros(m, dtype=bool)
            # For each keyword, add the file path to the string of files that contain it
            for keyword in keywords:
                #print(keyword)
                kval=keyword[0]
                for i in range(1,k+1):                  
                    kval,index=computeIntHash(kval, m)
                    bf[index]=1
                    #print(f"keyword: f{keyword[0]}, index: {index}")
            #print(bf)
            bf_str = ' '.join(map(str, bf.astype(int)))
            e=encryptData(file_path,Kgroup)
            sheet[f'A{idx}'] = file_path
            sheet[f'B{idx}'] = bf_str
            sheet[f'C{idx}'] = e
            sheet[f'D{idx}'] = k
            sheet[f'E{idx}'] = m
            idx=idx+1
            #print(e)
    wb.save('PCloudDataBase.xlsx')              
    print("Data has been saved")  
                


def token_Genration(searchQuery):
    token = hashlib.sha256(searchQuery.encode()).hexdigest()
    return token

def serach_Query(token,KGroup,file_path):
    wb = openpyxl.load_workbook(file_path)

    sheet = wb.active

    # Initialize a list to store the data retrieved from the Excel file
    retrieved_data = []
    pt=""
    # Iterate over rows starting from the second row (ignoring the header)
    for row in sheet.iter_rows(min_row=2, values_only=True):
        # Read data from columns A, B, C, and D
        path = row[0]
        bf_str = row[1]
        e = row[2]
        k = row[3]
        m = row[4]
        # Convert the Bloom filter string back to a NumPy array of booleans
        bf = np.array(list(map(int, bf_str.split())), dtype=bool)
        #szz=len(bf)
        #print(f"m:{m}, bfsize:{szz}")
        #print(len(bf))
        count=0
        #print(m)
        tcopy=token
        findex= int(tcopy, 16)%m
        if findex<len(bf):
            if bf[findex]==1:
                #print(f"{path} first {findex}")
                count+=1 
            for i in range(1,k):                  
                kval,rindex=computeIntHash(tcopy, m)
                tcopy=kval
                #print(f"{path} r index {rindex}")
                if(rindex<len(bf)):
                    if bf[rindex]==1:
                        count+=1  
        if(count==k):
            pt=pt+e+" "
        print(f"{path}, count: {count}, k={k}")           
        # Process the retrieved data as needed
        # You can process 'path', 'bf', 'e', and 'k' here
    return pt  



def findBloomsParas(n):
    # n is total keyword, #p fase positive rate 1%, #m bloom filter size, k number of hashs
    p=0.00001
    v1=-n * math.log(p)
    v2=math.log(2) ** 2
    print(f"v1: {v1}, v2: {v2}")    
    m = round(v1/v2)
    k = round(m / n * math.log(2))
    #print(f"ln: {math.log(p)} {math.log(2)}")
    return k, m, p


#Main function


dataset_path = r"maildir"  # Replace with your dataset path
cloud_path=r"PCloudDataBase.xlsx"
KGroup="AFDDDFFFEG";
#kw="Gdsfafeerf";
setup_Phase(dataset_path,KGroup)




#Token Generation Phase
squery="multexinvestor"
#squery="fdgasgdsdsafammemmemk44afasdfasd"
token=token_Genration(squery)







#Search Query Phase
pt=serach_Query(token,KGroup,cloud_path)

split_string = pt.split(' ')

# Iterate over each word
for p in split_string[:-1]:
    fid=decryptData(p, KGroup)
    print(fid)
# kpri,t,f=serach_Query(wt)
#n=198
#k,m,p=findBloomsParas(n)
#print(f"n: {n}, k: {k}, m: {m}, p: {p}")